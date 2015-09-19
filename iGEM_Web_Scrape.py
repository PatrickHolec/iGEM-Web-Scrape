
'''
_________ _______  _______  _______             _______  ______     _______  _______  _______  _______  _______  _______  _______ 
\__   __/(  ____ \(  ____ \(       )  |\     /|(  ____ \(  ___ \   (  ____ \(  ____ \(  ____ )(  ___  )(  ____ )(  ____ \(  ____ )
   ) (   | (    \/| (    \/| () () |  | )   ( || (    \/| (   ) )  | (    \/| (    \/| (    )|| (   ) || (    )|| (    \/| (    )|
   | |   | |      | (__    | || || |  | | _ | || (__    | (__/ /   | (_____ | |      | (____)|| (___) || (____)|| (__    | (____)|
   | |   | | ____ |  __)   | |(_)| |  | |( )| ||  __)   |  __ (    (_____  )| |      |     __)|  ___  ||  _____)|  __)   |     __)
   | |   | | \_  )| (      | |   | |  | || || || (      | (  \ \         ) || |      | (\ (   | (   ) || (      | (      | (\ (   
___) (___| (___) || (____/\| )   ( |  | () () || (____/\| )___) )  /\____) || (____/\| ) \ \__| )   ( || )      | (____/\| ) \ \__
\_______/(_______)(_______/|/     \|  (_______)(_______/|/ \___/   \_______)(_______/|/   \__/|/     \||/       (_______/|/   \__/

'''


'''

This collection of functions is designed to pull information out of websites and do some fundamental sentiment analyses and
linguistic analysis of content. This can also pull out information of temporal ranges to give you a feel for how people are talking
about a certain query over a timespan.

If you have any questions, comments, or want to talk about something interesting, please contact:

Patrick V. Holec
hole0077@umn.edu
University of Minnesota

'''



# Standard Libaries
import jdcal
import datetime
import urllib2
import collections
import urllib
import re,os,csv,time
from nltk.tag import pos_tag
import math

# Non-Standard Libaries
from BeautifulSoup import BeautifulSoup
import bs4
import google
from google import search
import xlrd
from openpyxl import Workbook
from textblob import TextBlob


##################################################
# Fill in any commands into the run section here #
##################################################



def Run():
    '''
    Fill in any combination of the three main function (Word_Scrape(),Google_Time_Lapse(), and Web_Scrape_Metrics()) to be executed
    You can choose some off shoot programs if you want but it'll probably be a little tricky unless you really understand the subprograms
    See each of these functions below for details on how each function works and what inputs are necessary. Good luck!
    '''
    pass
    
##################################################
#                                                #
##################################################


'''
Purpose: Word Puddle from Query
Descriptions: Pulls top articles, melts content to words, great for frequency analysis or word clouds
'''
def Word_Scrape(key_phrase,fname='Search_results'):
    PRO = WebScrape(key_phrase)
    temp = [item for sublist in PRO.words for item in sublist]
    Publish_TXT(fname+'.txt',' '.join(temp))
    print 'Results saved in ',fname+'.txt!'
    
'''
Purpose: AFINN Conversion
Descriptions: Opens AFINN lexicon and maps words to emotional magnitudes
'''   
def Google_Time_Lapse(key_phrase,date1=[2013,1],date2=[2015,1],sites_per_month=5):
    subjectivity,polarity = [],[]
    if date1[0] == date2[0]:  # If we're working in a daterange within a year
        i = date1[0]
        for j in xrange(date1[1],date2[1]+1):
            subjectivity,polarity = Google_Daterange(key_phrase,i,j,sites_per_month,subjectivity,polarity)
            subjectivity.append(temp[0])
            polarity.append(temp[1])

    else:  # If we're working in a daterange across multiple years
        i = date1[0]
        for j in xrange(date1[1],12+1):
            subjectivity,polarity = Google_Daterange(key_phrase,i,j,sites_per_month,subjectivity,polarity)
            subjectivity.append(temp[0])
            polarity.append(temp[1])
        for i in xrange(date1[0]+1,date1[1]):
            for j in xrange(date1[1],12+1):
                subjectivity,polarity = Google_Daterange(key_phrase,i,j,sites_per_month,subjectivity,polarity)
                subjectivity.append(temp[0])
                polarity.append(temp[1])
        i = date2[1]
        for j in xrange(1,date2[1]+1):
            subjectivity,polarity = Google_Daterange(key_phrase,i,j,sites_per_month,subjectivity,polarity)
            subjectivity.append(temp[0])
            polarity.append(temp[1])

    GeneralExcel(subjectivity,'Subjectivity of '+key_phrase+'.xlsx')  # Publishing data on subjectivity
    GeneralExcel(polarity,'Polarity '+key_phrase+'.xlsx')   # Publishing data on polarity


'''
Purpose: AFINN Conversion
Descriptions: Opens AFINN lexicon and maps words to emotional magnitudes
'''
def Web_Scape_Metrics(key_phrase,fname='Search_metrics'):
    data = WebScrape(key_phrase)
    Popular_Titles(data,fname+'.xlsx')
    print 'Results saved in ',fname+'.xlsx!'


###########################################



'''
Purpose: AFINN Conversion
Descriptions: Opens AFINN lexicon and maps words to emotional magnitudes
'''
def Google_Daterange(string,i,j,total,subjectivity,polarity):
    query = string+' daterange:'+JulianDate(j,i,1)+'-'+JulianDate(j,i,30)
    print 'Searching for:',query
    WEB = WebScrape(query,total)
    subjectivity.append(['Month '+str(j)+', Year '+str(i)])
    polarity.append(['Month '+str(j)+', Year '+str(i)])
    for words in WEB.raw_words:
        test = TextBlob(words)
        subjectivity[-1].append(test.sentiment.subjectivity)
        polarity[-1].append(test.sentiment.polarity)
    return subjectivity,polarity

'''
Purpose: Julian data converter
Descriptions: Google only interprets dates as Julian numbers, converts regular days to this for
'''
def JulianDate(month,year,day=1):
    return str(int(sum(jdcal.gcal2jd(year,month,day))))


'''
Purpose: Publish Text in Excel
Descriptions: See purpose
'''
def GeneralExcel(data,fname):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = datetime.datetime.now()
    for d in data:
        ws.append(d)
    wb.save(fname)


'''
Purpose: Dictionary Initialization
Descriptions: Creates dictionary for part of speech separation with nltk module
'''
def Dictionaries():
    global tags, AFINN_key, AFINN_value
    tags = [('CC','coordinating conjunction'),
    ('CD','cardinal number'),
    ('DT','determiner'),
    ('EX','	existential there'),
    ('FW','	foreign word'),
    ('IN','	preposition/subordinating conjunction'),
    ('JJ','	adjective'),
    ('JJR','adjective, comparative'),
    ('JJS','adjective, superlative'),
    ('LS','	list marker'),
    ('MD','	modal'),
    ('NN','	noun, singular or mass'),
    ('NNS','noun plural'),
    ('NNP','proper noun, singular'),
    ('NNPS','proper noun, plural'),
    ('PDT','predeterminer'),
    ('POS','possessive ending'),
    ('PRP','personal pronoun'),
    ('PRP$','possessive pronoun'),
    ('RB','adverb'),
    ('RBR','adverb, comparative'),
    ('RBS','adverb, superlative'),
    ('RP ','particle'),
    ('TO ','to'),
    ('UH ','interjection'),
    ('VB ','verb, base form'),
    ('VBD','verb, past tense'),
    ('VBG','verb, gerund/present participle'),
    ('VBN','verb, past participle'),
    ('VBP','verb, sing. present, non-3d'),
    ('VBZ','verb, 3rd person sing. present'),
    ('WDT','wh-determiner'),
    ('WP ','wh-pronoun'),
    ('WP$','possessive wh-pronoun'),
    ('WRB','wh-abverb')]
    with open('AFINN.csv', 'rb') as f:
        reader = csv.reader(f)
        AFINN = list(reader)
    AFINN_key = [i[0].replace('\t','') for i in AFINN]
    AFINN_value = [int(i[1]) for i in AFINN]


'''
Purpose: AFINN Conversion
Descriptions: Opens AFINN lexicon and maps words to emotional magnitudes
'''
def AFINN_Analysis(text):
    text = ' '.join(text)
    text = text.lower().split(' ')
    key = [-5,-4,-3,-2,-1,0,1,2,3,4,5]
    value = [0,0,0,0,0,0,0,0,0,0,0]
    text = [i.encode('ascii','ignore') for i in text]
    for word in text:
        if word in AFINN_key:
            value[key.index(AFINN_value[AFINN_key.index(word)])] += 1
    return value,len(text)

'''
Purpose: Identifies Words in Text
Descriptions: Breaks up input text into words
'''
def Word_ID(background):
    WORDS,IDS,text = [],[],[]
    for title in [i for i in background]:
        try:
            tagged_title = pos_tag(title.split())
            words,ids = [i[0] for i in tagged_title],[i[1] for i in tagged_title]
            IDS += ids
            text.append(title)
        except:
            print 'Title rejected due to unknown data format.'
    counter=collections.Counter(IDS)
    return counter.keys(),counter.values(),text

def Word_ID2(background):
    WORDS,IDS,text = [],[],[]
    for title in [i for i in background]:
        try:
            tagged_title = pos_tag(title.split())
            words += [i[0] for i in tagged_title],[i[1] for i in tagged_title]
            IDS += ids
            text.append(title)
        except:
            print 'Title rejected due to unknown data format.'
    counter=collections.Counter(IDS)
    return counter.keys(),counter.values(),text

'''
Purpose: Analyzes Text for Metrics
Descriptions: Iterates through lists of text for analysis
'''
def Popular_Titles(background,fname,iteration = 1):
    iteration = 1   # Number of titles are binned together for analysis
    wb = Workbook()
    ws = wb.active

    spectrum = ['-5','-4','-3','-2','-1','0','1','2','3','4','5']
    angles = ['','','']
    
    ws['A1'] = datetime.datetime.now()
    ws.append(['','','Correlation:']+[i[0] for i in tags])
    ws.append(['Article Range']+[i[1] for i in tags]+['Net Emotion','Emotional Magnitude','Subjectivity','Polarity','Words/Title','','All words'])
    results,pca_data = [],[]
    for stack in range(0,len(background),iteration):
        print 'Starting analysis on titles:',str(stack+1)+'-'+str(stack+iteration)
        if stack + iteration - 1 < len(background) - 1:
            results.append([str(stack+1)+'-'+str(stack+iteration)])
        else:
            results.append([str(stack+1)+'-'+str(len(background))])
        bg = background[stack:stack+iteration]
        keys,values,text = Word_ID(bg)
        emo,count = AFINN_Analysis(text)
        for tag,temp in tags:
            if tag in keys:
                results[-1].append(float(values[keys.index(tag)])/count)
            else:
                results[-1].append(0)
        words = TextBlob('. '.join(text))
        results[-1] += [float(sum([int(i)*j for i,j in zip(spectrum,emo)]))/iteration,float(sum([abs(int(i))*j for i,j in zip(spectrum,emo)]))/iteration,words.sentiment.subjectivity,words.sentiment.polarity,float(count)/iteration]
        results[-1] += ['','. '.join(text)]
    for i in results:
        ws.append(i)
    wb.save(fname)
    print 'Analysis completed!'

    
'''
Purpose: Google Search URLs
Descriptions: Returns "total" (int) number of URLs in response to a google search of "query" (string)
Powered by Google!
'''

def SearchResults(query,total):
    results = []
    for url in google.search(query, num=total, stop=1):
        results.append(url)
    return results


'''
Purpose: Core Web Scraper
Descriptions: Pulls information off websites and puts it into useful forms
Class structured for accessibility
'''

class WebScrape:
    '''
    Purpose: Initializes Class
    Descriptions: Prepares the web scraping string for analysis
    '''
    def __init__(self,query,article_total=10):
        self.article_total = article_total
        self.nbhd = 5
        check = False
        
        self.urls = SearchResults(query,self.article_total)
        self.words,self.hs_words,self.titles,self.raw_words = [],[],[],[]
        
        hotspots = [['gmo'],['genetically','modified','organism'],['genetically','modified','food']]
        for i,url in enumerate(self.urls):
            print 'Loading site: %d \n URL: %s' % (i,url)
            filename,extension = os.path.splitext('/path/to/somefile.ext')
            if extension == '.pdf':
                print 'PDF file rejected due to formatting problems.'
            else:
                self.URL_Format(url,i)
        if check:
            keepers = []
            for title in self.words:
                try:
                    temp = raw_input('Rename: ')
                    if len(temp) > 0:
                        keepers.append([temp,title[1]])
                        print 'Saved!'
                    else:
                        print 'Rejected.'
                except:
                    print 'That was weird. Skipping entry.'
            Popular_Titles(keepers,query+'.xlsx')
        else:
            Popular_Titles(self.words,query+'.xlsx')
        self.Hotspot_Analysis(self.words[-1],hotspots)

    '''
    Purpose: URL to HTML Visible Text
    Descriptions: Converts a website indexed by a URL into a body of text based on what is visible
    Critical to webscaping
    '''

    def URL_Format(self,url,index):  
        try:
            opener = urllib2.build_opener()
            print 'Opener.'
            opener.addheaders = [('User-agent','Google Chrome')]
            print 'Add headers.'
            response = opener.open(url)
            print 'Open URL.'
            page = response.read()
            print 'Response.'
            soup = BeautifulSoup(page)
            print 'Formatting webppage titled: %s' % soup.title.string
            texts = soup.findAll(text=True)
            visible_texts = filter(visible, texts)
            visible_texts = ' '.join(visible_texts)
            visible_texts = visible_texts.replace('\n',' ').replace('\t',' ')
            self.raw_words.append(visible_texts)
            
            temp = ' '.join(visible_texts.split())
            temp = Remove_Characters(temp,"""#@$.?(),'":;^[]""")
            temp = [word.lower() for word in temp.split(' ')]
            words = [x for x in temp if not any(c.isdigit() for c in x)]
            self.titles.append([soup.title.string,index])
            self.words.append(words)
        except:
            print 'This website is acting very strange, entry rejected.'

    '''
    Purpose: Hotspot Analysis
    Descriptions: Finds a collections of words surround hotspot words in a body of text
    '''
    def Hotspot_Analysis(self,words,hotspots):
        for hotspot in hotspots:
            lp = len(hotspot)
            self.hs_words += list(words[j+i] for j in (range(-self.nbhd,0) + range(lp,self.nbhd+lp)) for i,x in enumerate(words[:-(lp+self.nbhd)]) if words[i:i+lp] == hotspot)
            #print 'Singular search:',len(self.hs_words)
            self.hs_words += list(words[j+i] for j in (range(-self.nbhd,0) + range(lp,self.nbhd+lp)) for i,x in enumerate(words[:-(lp+self.nbhd)]) if words[i:i+lp] == hotspot[:-1]+[hotspot[-1]+'s'])
            #print 'Plural search:',len(self.hs_words)

'''
Purpose: Visible Text Filter
Descriptions: Filters html to check for visibility
Converts raw html to something useful
'''
def visible(element):
    if element.parent.name in ['style', 'script', '[document]', 'head', 'title']:
        return False
    elif re.match('<!--.*-->', str(element)):
        return False
    return True

'''
Purpose: Removes Characters
Descriptions: Filters out odd characters from string to prevent algorithm from getting confused
'''
def Remove_Characters(string,chars):
    for char in chars:
        string = string.replace(char,'')
    return string

'''
Purpose: Publishes Results
Descriptions: Publishes text to txt file
'''
def Publish_TXT(fname,data):
    with open(fname, 'w') as text_file:
        text_file.write(data.encode('utf8'))

Dictionaries()
Run()
